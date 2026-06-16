import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

DEFAULT_FILENAME = "Internship_Log.xlsx"

def apply_borders_to_range(ws, start_row, start_col, end_row, end_col):
    thin_side = Side(border_style="thin", color="BFBFBF")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border

def initialize_timesheet_sheet(ws):
    ws.views.sheetView[0].showGridLines = True
    
    # Title in A1:F1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "INTERNSHIP - TIMESHEET LOG"
    title_cell.font = Font(name="Aptos Narrow", size=12, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    title_cell.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 26
    
    # Headers in A2:F2
    headers = ["Date", "Start Time", "End Time", "Duration (Hrs)", "Category", "Activity Log"]
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    header_font = Font(name="Aptos Narrow", size=11, bold=True, color="000000")
    header_align = Alignment(vertical="center", horizontal="center")
    
    ws.row_dimensions[2].height = 22
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        
    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 60
    
    apply_borders_to_range(ws, 1, 1, 2, 6)

def get_timesheet_sheet(filepath):
    dir_name = os.path.dirname(filepath)
    if dir_name:
        os.makedirs(dir_name, exist_ok=True)
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Timesheet"
        initialize_timesheet_sheet(ws1)
        
    if "Timesheet" not in wb.sheetnames:
        ws = wb.create_sheet(title="Timesheet")
        initialize_timesheet_sheet(ws)
    else:
        ws = wb["Timesheet"]
        
    wb.active = ws
    return wb, ws

def generate_timesheet_slots(arrival_str, work_hours=8, interval_hours=2, lunch_start_str="13:00", lunch_end_str="14:00"):
    try:
        arrival_t = datetime.strptime(arrival_str.strip(), "%H:%M")
    except ValueError:
        arrival_t = datetime.strptime("09:00", "%H:%M")
        
    lunch_start_t = datetime.strptime(lunch_start_str, "%H:%M")
    lunch_end_t = datetime.strptime(lunch_end_str, "%H:%M")
    
    start_min = arrival_t.hour * 60 + arrival_t.minute
    lunch_start = lunch_start_t.hour * 60 + lunch_start_t.minute
    lunch_end = lunch_end_t.hour * 60 + lunch_end_t.minute
    lunch_duration = lunch_end - lunch_start
    
    total_work_min = work_hours * 60
    interval_min = interval_hours * 60
    
    curr = start_min
    allocated_work = 0
    slots = []
    
    while allocated_work < total_work_min:
        remaining_block_work = min(interval_min, total_work_min - allocated_work)
        
        if curr < lunch_start:
            work_before_lunch = lunch_start - curr
            if work_before_lunch >= remaining_block_work:
                slots.append({"start": curr, "end": curr + remaining_block_work, "type": "Work", "duration": remaining_block_work / 60.0})
                allocated_work += remaining_block_work
                curr += remaining_block_work
            else:
                if work_before_lunch > 0:
                    slots.append({"start": curr, "end": lunch_start, "type": "Work", "duration": work_before_lunch / 60.0})
                    allocated_work += work_before_lunch
                slots.append({"start": lunch_start, "end": lunch_end, "type": "Lunch Break", "duration": lunch_duration / 60.0})
                rem = remaining_block_work - work_before_lunch
                slots.append({"start": lunch_end, "end": lunch_end + rem, "type": "Work", "duration": rem / 60.0})
                allocated_work += rem
                curr = lunch_end + rem
        elif curr >= lunch_start and curr < lunch_end:
            lunch_left = lunch_end - curr
            slots.append({"start": curr, "end": lunch_end, "type": "Lunch Break", "duration": lunch_left / 60.0})
            curr = lunch_end
        else:
            slots.append({"start": curr, "end": curr + remaining_block_work, "type": "Work", "duration": remaining_block_work / 60.0})
            allocated_work += remaining_block_work
            curr += remaining_block_work
            
    formatted_slots = []
    for slot in slots:
        sh = int(slot["start"] // 60)
        sm = int(slot["start"] % 60)
        eh = int(slot["end"] // 60)
        em = int(slot["end"] % 60)
        formatted_slots.append({
            "start": f"{sh:02d}:{sm:02d}",
            "end": f"{eh:02d}:{em:02d}",
            "type": slot["type"],
            "duration": slot["duration"]
        })
    return formatted_slots

def find_date_row_range(ws, date_str):
    for r in range(3, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == date_str:
            start_row = r
            end_row = r
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_row == r and merged_range.min_col == 1:
                    end_row = merged_range.max_row
                    break
            return start_row, end_row
    return None, None

def get_next_available_row(ws):
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, 7)):
            return r + 1
    return 3

def get_or_create_day_slots(filepath, date_str, arrival_time):
    wb, ws = get_timesheet_sheet(filepath)
    start_row, end_row = find_date_row_range(ws, date_str)
    
    if start_row is not None:
        slots = []
        for r in range(start_row, end_row + 1):
            slots.append({
                "row": r,
                "start": ws.cell(row=r, column=2).value,
                "end": ws.cell(row=r, column=3).value,
                "duration": ws.cell(row=r, column=4).value,
                "type": ws.cell(row=r, column=5).value,
                "activity": ws.cell(row=r, column=6).value
            })
        wb.close()
        return slots
    
    raw_slots = generate_timesheet_slots(arrival_time)
    start_row = get_next_available_row(ws)
    end_row = start_row + len(raw_slots) - 1
    
    ws.cell(row=start_row, column=1, value=date_str)
    
    for i, slot in enumerate(raw_slots):
        r = start_row + i
        ws.cell(row=r, column=2, value=slot["start"])
        ws.cell(row=r, column=3, value=slot["end"])
        ws.cell(row=r, column=4, value=slot["duration"])
        ws.cell(row=r, column=5, value=slot["type"])
        
        if slot["type"] == "Lunch Break":
            ws.cell(row=r, column=6, value="Lunch Break")
        else:
            ws.cell(row=r, column=6, value="")
            
    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
    
    thin_side = Side(border_style="thin", color="BFBFBF")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    lunch_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for r in range(start_row, end_row + 1):
        is_lunch = (ws.cell(row=r, column=5).value == "Lunch Break")
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = Font(name="Aptos Narrow", size=11, bold=(c == 1))
            
            if c in (1, 2, 3, 4, 5):
                cell.alignment = Alignment(vertical="center", horizontal="center")
            else:
                cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
                
            if is_lunch:
                cell.fill = lunch_fill
                
    wb.save(filepath)
    wb.close()
    
    return get_or_create_day_slots(filepath, date_str, arrival_time)

def recreate_day_slots(filepath, date_str, arrival_time):
    wb, ws = get_timesheet_sheet(filepath)
    start_row, end_row = find_date_row_range(ws, date_str)
    
    if start_row is not None:
        try:
            ws.unmerge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        except Exception:
            pass
            
        ws.delete_rows(start_row, amount=(end_row - start_row + 1))
        
    wb.save(filepath)
    wb.close()
    return get_or_create_day_slots(filepath, date_str, arrival_time)

def save_timesheet_slot_activity(filepath, row_idx, activity_text):
    wb = load_workbook(filepath)
    if "Timesheet" not in wb.sheetnames:
        wb.close()
        raise ValueError("Timesheet sheet does not exist.")
    ws = wb["Timesheet"]
    
    cell = ws.cell(row=row_idx, column=6)
    cell.value = activity_text
    cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    cell.font = Font(name="Aptos Narrow", size=11)
    
    wb.active = ws
    wb.save(filepath)
    wb.close()

def get_activities_for_date(filepath, date_str):
    wb = load_workbook(filepath)
    if "Timesheet" not in wb.sheetnames:
        wb.close()
        return []
    ws_ts = wb["Timesheet"]
    
    start_row, end_row = find_date_row_range(ws_ts, date_str)
    if not start_row:
        wb.close()
        return []
        
    activities = []
    for r in range(start_row, end_row + 1):
        cat = ws_ts.cell(row=r, column=5).value
        act = ws_ts.cell(row=r, column=6).value
        if cat == "Work" and act:
            act_str = str(act).strip()
            if act_str:
                activities.append(act_str)
                
    wb.close()
    return activities
