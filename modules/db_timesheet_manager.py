import os
from datetime import datetime
import user_manager
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def generate_timesheet_slots(arrival_str, work_hours=8, interval_hours=2, lunch_start_str="13:00", lunch_end_str="14:00"):
    try:
        arrival_t = datetime.strptime(arrival_str.strip(), "%H:%M")
    except ValueError:
        arrival_t = datetime.strptime("09:00", "%H:%M")
        
    lunch_start_t = datetime.strptime(lunch_start_str, "%H:%M")
    lunch_end_t = datetime.strptime(lunch_end_str, "%H:%M")
    
    start_min = arrival_t.hour * 60 + arrival_t.minute
    lunch_start = lunch_start_t.hour * 60 + lunch_start_t.minute
    lunch_end = lunch_end_t.hour * 60 + lunch_end_t.minute
    lunch_duration = lunch_end - lunch_start
    
    total_work_min = work_hours * 60
    interval_min = interval_hours * 60
    
    curr = start_min
    allocated_work = 0
    slots = []
    
    while allocated_work < total_work_min:
        remaining_block_work = min(interval_min, total_work_min - allocated_work)
        
        if curr < lunch_start:
            work_before_lunch = lunch_start - curr
            if work_before_lunch >= remaining_block_work:
                slots.append({"start": curr, "end": curr + remaining_block_work, "type": "Work", "duration": remaining_block_work / 60.0})
                allocated_work += remaining_block_work
                curr += remaining_block_work
            else:
                if work_before_lunch > 0:
                    slots.append({"start": curr, "end": lunch_start, "type": "Work", "duration": work_before_lunch / 60.0})
                    allocated_work += work_before_lunch
                slots.append({"start": lunch_start, "end": lunch_end, "type": "Lunch Break", "duration": lunch_duration / 60.0})
                rem = remaining_block_work - work_before_lunch
                slots.append({"start": lunch_end, "end": lunch_end + rem, "type": "Work", "duration": rem / 60.0})
                allocated_work += rem
                curr = lunch_end + rem
        elif curr >= lunch_start and curr < lunch_end:
            lunch_left = lunch_end - curr
            slots.append({"start": curr, "end": lunch_end, "type": "Lunch Break", "duration": lunch_left / 60.0})
            curr = lunch_end
        else:
            slots.append({"start": curr, "end": curr + remaining_block_work, "type": "Work", "duration": remaining_block_work / 60.0})
            allocated_work += remaining_block_work
            curr += remaining_block_work
            
    formatted_slots = []
    for slot in slots:
        sh = int(slot["start"] // 60)
        sm = int(slot["start"] % 60)
        eh = int(slot["end"] // 60)
        em = int(slot["end"] % 60)
        formatted_slots.append({
            "start": f"{sh:02d}:{sm:02d}",
            "end": f"{eh:02d}:{em:02d}",
            "type": slot["type"],
            "duration": f"{slot['duration']:.2f}"
        })
    return formatted_slots

def get_or_create_day_slots(username, date_val, arrival_time):
    username = username.strip().lower()
    conn = user_manager.get_db()
    
    cursor = conn.execute("SELECT * FROM timesheet_slots WHERE username = ? AND date_val = ? ORDER BY row_index ASC", (username, date_val))
    rows = cursor.fetchall()
    
    if rows:
        slots = []
        for r in rows:
            slots.append({
                "row": r["row_index"],
                "start": r["start_time"],
                "end": r["end_time"],
                "duration": r["duration_hrs"],
                "type": r["category"],
                "activity": r["activity_text"]
            })
        conn.close()
        return slots
        
    # Generate new slots
    raw_slots = generate_timesheet_slots(arrival_time)
    slots = []
    for i, slot in enumerate(raw_slots):
        row_index = i + 1
        activity = "Lunch Break" if slot["type"] == "Lunch Break" else ""
        conn.execute(
            "INSERT INTO timesheet_slots (username, date_val, start_time, end_time, duration_hrs, category, activity_text, row_index) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (username, date_val, slot["start"], slot["end"], slot["duration"], slot["type"], activity, row_index)
        )
        slots.append({
            "row": row_index,
            "start": slot["start"],
            "end": slot["end"],
            "duration": slot["duration"],
            "type": slot["type"],
            "activity": activity
        })
        
    conn.commit()
    conn.close()
    return slots

def save_timesheet_slot_activity(username, date_val, row_idx, activity_text):
    username = username.strip().lower()
    conn = user_manager.get_db()
    conn.execute(
        "UPDATE timesheet_slots SET activity_text = ? WHERE username = ? AND date_val = ? AND row_index = ?",
        (activity_text, username, date_val, row_idx)
    )
    conn.commit()
    conn.close()

def get_day_activities(username, date_val):
    username = username.strip().lower()
    conn = user_manager.get_db()
    cursor = conn.execute(
        "SELECT start_time, end_time, activity_text FROM timesheet_slots WHERE username = ? AND date_val = ? AND category != 'Lunch Break' ORDER BY row_index ASC",
        (username, date_val)
    )
    rows = cursor.fetchall()
    conn.close()
    
    activities = []
    for r in rows:
        if r["activity_text"] and r["activity_text"].strip():
            activities.append(f"{r['start_time']} to {r['end_time']}: {r['activity_text']}")
    return activities

def apply_borders_to_range(ws, start_row, start_col, end_row, end_col):
    thin_side = Side(border_style="thin", color="BFBFBF")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border

def generate_excel_download(username):
    username = username.strip().lower()
    conn = user_manager.get_db()
    cursor = conn.execute("SELECT * FROM timesheet_slots WHERE username = ? ORDER BY date_val ASC, row_index ASC", (username,))
    rows = cursor.fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    ws.views.sheetView[0].showGridLines = True
    
    # Title in A1:F1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "INTERNSHIP - TIMESHEET LOG"
    title_cell.font = Font(name="Aptos Narrow", size=12, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    title_cell.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 26
    
    # Headers in A2:F2
    headers = ["Date", "Start Time", "End Time", "Duration (Hrs)", "Category", "Activity Log"]
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    header_font = Font(name="Aptos Narrow", size=11, bold=True, color="000000")
    header_align = Alignment(vertical="center", horizontal="center")
    
    ws.row_dimensions[2].height = 22
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        
    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 60
    
    apply_borders_to_range(ws, 1, 1, 2, 6)
    
    if not rows:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
        
    current_row = 3
    thin_side = Side(border_style="thin", color="BFBFBF")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    lunch_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Group by date
    grouped = {}
    for r in rows:
        date_val = r["date_val"]
        if date_val not in grouped:
            grouped[date_val] = []
        grouped[date_val].append(r)
        
    for date_val, slots in grouped.items():
        start_row = current_row
        
        for slot in slots:
            is_lunch = (slot["category"] == "Lunch Break")
            
            ws.cell(row=current_row, column=1, value=date_val)
            ws.cell(row=current_row, column=2, value=slot["start_time"])
            ws.cell(row=current_row, column=3, value=slot["end_time"])
            ws.cell(row=current_row, column=4, value=slot["duration_hrs"])
            ws.cell(row=current_row, column=5, value=slot["category"])
            ws.cell(row=current_row, column=6, value=slot["activity_text"])
            
            for c in range(1, 7):
                cell = ws.cell(row=current_row, column=c)
                cell.border = thin_border
                cell.font = Font(name="Aptos Narrow", size=11, bold=(c == 1))
                
                if c in (1, 2, 3, 4, 5):
                    cell.alignment = Alignment(vertical="center", horizontal="center")
                else:
                    cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
                    
                if is_lunch:
                    cell.fill = lunch_fill
                    
            current_row += 1
            
        end_row = current_row - 1
        if start_row < end_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
